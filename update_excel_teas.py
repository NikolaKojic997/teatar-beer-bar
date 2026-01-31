
import openpyxl
import os

input_file = 'public/menu.xlsx'

if not os.path.exists(input_file):
    print(f"File {input_file} not found.")
    exit(1)

try:
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # Find columns
    headers = [cell.value for cell in sheet[1]]
    try:
        naziv_col = headers.index('Naziv') + 1
        grupa_col = headers.index('Grupa') + 1
    except ValueError:
        print("Required columns 'Naziv' or 'Grupa' not found.")
        exit(1)

    tea_keywords = ['NANA', 'KAMILICA', 'ZELENI', 'VOĆNI', 'VOCNI', 'CRNI', 'BRUSNICA', 'HIBISKUS', 'ŠUMSKO VOĆE', 'SUMSKO VOCE', 'PLANINSKI', 'ŠIPAK', 'SIPAK', 'VITAMINSKI']
    
    updated_count = 0
    for row in range(2, sheet.max_row + 1):
        naziv = sheet.cell(row=row, column=naziv_col).value
        grupa = sheet.cell(row=row, column=grupa_col).value
        
        if naziv and grupa and str(grupa).upper() == 'KAFE I ČAJEVI':
            naziv_str = str(naziv).strip()
            naziv_upper = naziv_str.upper()
            
            # Check if it's likely a tea and doesn't already have 'ČAJ'
            is_tea = any(kw in naziv_upper for kw in tea_keywords)
            has_caj = 'ČAJ' in naziv_upper or 'CAJ' in naziv_upper
            
            if is_tea and not has_caj:
                sheet.cell(row=row, column=naziv_col).value = f"ČAJ {naziv_str}"
                updated_count += 1
                print(f"Updated: {naziv_str} -> ČAJ {naziv_str}")

    wb.save(input_file)
    print(f"Successfully updated {updated_count} items in Excel.")

except Exception as e:
    print(f"Error: {e}")
    exit(1)
