
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import sys
import os

# Define input and output paths
input_file = 'public/menu.xlsx'
output_file = 'public/menu.xlsx'

try:
    print(f"Reading {input_file}...")
    # Try reading .xls first, if it fails try .xlsx
    if os.path.exists(input_file):
        df = pd.read_excel(input_file)
    elif os.path.exists(output_file):
        df = pd.read_excel(output_file)
    else:
        print("No menu file found.")
        sys.exit(1)
    
    # 1. Standardize 'Tip'
    if 'Tip' in df.columns:
        print("Standardizing 'Tip' column...")
        df['Tip'] = df['Tip'].astype(str).str.strip().str.title()
        
    # 2. Merge Wine Categories in 'Grupa'
    if 'Grupa' in df.columns:
        print("Merging Wine groups...")
        df['Grupa'] = df['Grupa'].astype(str).str.strip()
        mask = df['Grupa'].str.upper() == 'VINA FLASICE'
        df.loc[mask, 'Grupa'] = 'VINA'
        
        # Get unique groups for dropdown
        unique_groups = sorted([g for g in df['Grupa'].unique() if pd.notna(g) and g != 'nan'])
        print("Unique Groups for dropdown:", unique_groups)

    print(f"Saving to {output_file} using openpyxl with data validation...")
    
    # Use pandas to write to excel first (helper)
    df.to_excel(output_file, index=False, engine='openpyxl')
    
    # Now open with openpyxl to add validation
    from openpyxl import load_workbook
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Create the validation object with strict error checking
    formula = f'"{",".join(unique_groups)}"'
    dv = DataValidation(
        type="list", 
        formula1=formula, 
        allow_blank=True,
        showErrorMessage=True,
        errorStyle='stop',
        errorTitle='Nevažeća grupa',
        error='Molimo izaberite grupu iz ponuđene liste.'
    )
    
    # Find the 'Grupa' column index
    grupa_col_idx = -1
    for cell in ws[1]:
        if cell.value == 'Grupa':
            grupa_col_idx = cell.column
            break
    
    if grupa_col_idx != -1:
        col_letter = get_column_letter(grupa_col_idx)
        # Apply to column (from row 2 to 1000)
        dv.add(f"{col_letter}2:{col_letter}1000")
        ws.add_data_validation(dv)
        print(f"Added dropdown validation to column {col_letter}")
    
    wb.save(output_file)
    print("Done.")

    # Remove old .xls if it exists and we have .xlsx now
    if os.path.exists(input_file) and input_file != output_file:
        # os.remove(input_file) # Optional: keeping it for now to be safe, or rename it
        pass

except Exception as e:
    print(f"Error processing file: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
